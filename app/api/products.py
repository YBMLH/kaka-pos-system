"""Product management API: CRUD, search, archive, duplicate, history, import/export."""
import io

from flask import Blueprint, jsonify, request, send_file

from ..database import execute, get_db, query
from ..utils.audit import log_action
from ..utils.security import login_required, permission_required
from ..utils.text import build_search_blob, levenshtein, normalize

bp = Blueprint("products", __name__, url_prefix="/api/products")

PRODUCT_COLUMNS = (
    "barcode", "sku", "name_en", "name_ar", "name_fr", "category_id", "brand_id",
    "supplier_id", "purchase_price", "selling_price", "tax_rate", "quantity",
    "min_stock", "unit", "image", "expiry_date", "batch_number", "notes",
)


def _serialize(row) -> dict:
    d = dict(row)
    pp = d.get("purchase_price") or 0
    sp = d.get("selling_price") or 0
    d["profit_margin"] = round(((sp - pp) / sp * 100), 2) if sp else 0
    d["profit_unit"] = round(sp - pp, 2)
    d["low_stock"] = (d.get("quantity") or 0) <= (d.get("min_stock") or 0)
    return d


def _clean(data: dict) -> dict:
    out = {}
    for col in PRODUCT_COLUMNS:
        if col in data:
            out[col] = data[col]
    for num in ("purchase_price", "selling_price", "tax_rate", "quantity", "min_stock"):
        if num in out and out[num] in ("", None):
            out[num] = 0
    for ref in ("category_id", "brand_id", "supplier_id"):
        if ref in out and out[ref] in ("", 0):
            out[ref] = None
    return out


@bp.get("")
@login_required
def list_products():
    args = request.args
    page = max(int(args.get("page", 1)), 1)
    per_page = min(int(args.get("per_page", 50)), 200)
    search = (args.get("q") or "").strip()
    category = args.get("category_id")
    archived = args.get("archived", "0")
    low_stock = args.get("low_stock")
    sort = args.get("sort", "name_en")
    order = "DESC" if args.get("order", "asc").lower() == "desc" else "ASC"

    allowed_sort = {"name_en", "selling_price", "quantity", "purchase_price", "created_at", "id"}
    if sort not in allowed_sort:
        sort = "name_en"

    where = ["p.is_archived = ?"]
    params = [1 if archived == "1" else 0]
    if search:
        norm = normalize(search)
        where.append("(p.search_blob LIKE ? OR p.barcode LIKE ? OR p.sku LIKE ?)")
        params += [f"%{norm}%", f"%{search}%", f"%{search}%"]
    if category:
        where.append("p.category_id = ?")
        params.append(category)
    if low_stock == "1":
        where.append("p.quantity <= p.min_stock")

    where_sql = " AND ".join(where)
    total = query(f"SELECT COUNT(*) AS c FROM products p WHERE {where_sql}", params, one=True)["c"]
    rows = query(
        f"SELECT p.*, c.name AS category_name, b.name AS brand_name, "
        f"s.company_name AS supplier_name "
        f"FROM products p "
        f"LEFT JOIN categories c ON c.id = p.category_id "
        f"LEFT JOIN brands b ON b.id = p.brand_id "
        f"LEFT JOIN suppliers s ON s.id = p.supplier_id "
        f"WHERE {where_sql} ORDER BY p.{sort} {order} LIMIT ? OFFSET ?",
        params + [per_page, (page - 1) * per_page],
    )
    return jsonify({
        "products": [_serialize(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    })


@bp.get("/search")
@login_required
def search_products():
    """Fast instant-search for the POS: barcode-first, then typo-tolerant name."""
    term = (request.args.get("q") or "").strip()
    limit = min(int(request.args.get("limit", 15)), 50)
    if not term:
        return jsonify({"products": []})

    # 1) exact barcode wins instantly.
    exact = query("SELECT * FROM products WHERE barcode = ? AND is_archived = 0", (term,), one=True)
    if exact:
        return jsonify({"products": [_serialize(exact)], "exact_barcode": True})

    norm = normalize(term)
    rows = query(
        "SELECT * FROM products WHERE is_archived = 0 AND "
        "(search_blob LIKE ? OR sku LIKE ? OR barcode LIKE ?) LIMIT 80",
        (f"%{norm}%", f"%{term}%", f"%{term}%"),
    )
    # Rank: prefix matches first, then edit distance for typo tolerance.
    def score(r):
        blob = r["search_blob"] or ""
        if blob.startswith(norm):
            return (0, 0)
        best = min((levenshtein(norm, tok) for tok in blob.split()), default=99)
        return (1, best)

    ranked = sorted(rows, key=score)[:limit]
    return jsonify({"products": [_serialize(r) for r in ranked]})


@bp.get("/<int:pid>")
@login_required
def get_product(pid):
    row = query("SELECT * FROM products WHERE id = ?", (pid,), one=True)
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify({"product": _serialize(row)})


@bp.get("/<int:pid>/history")
@login_required
def product_history(pid):
    rows = query(
        "SELECT * FROM inventory_movements WHERE product_id = ? ORDER BY created_at DESC LIMIT 200",
        (pid,),
    )
    return jsonify({"movements": [dict(r) for r in rows]})


@bp.post("")
@permission_required("products.edit")
def create_product():
    data = request.get_json(silent=True) or {}
    fields = _clean(data)
    if not (fields.get("name_en") or fields.get("name_ar") or fields.get("name_fr")):
        return jsonify({"error": "at least one product name is required"}), 400
    barcode = fields.get("barcode")
    if barcode:
        dup = query("SELECT id FROM products WHERE barcode = ?", (barcode,), one=True)
        if dup:
            return jsonify({"error": "barcode already exists", "product_id": dup["id"]}), 409

    fields["search_blob"] = build_search_blob(
        fields.get("name_en", ""), fields.get("name_ar", ""),
        fields.get("name_fr", ""), fields.get("sku", ""), barcode or "",
    )
    cols = list(fields.keys())
    placeholders = ",".join("?" for _ in cols)
    pid = execute(
        f"INSERT INTO products ({','.join(cols)}) VALUES ({placeholders})",
        [fields[c] for c in cols],
    )
    qty = float(fields.get("quantity") or 0)
    if qty:
        _record_movement(pid, qty, qty, "initial", note="Initial stock")
    log_action("create_product", "product", pid, fields.get("name_en", ""))
    return jsonify({"id": pid}), 201


@bp.put("/<int:pid>")
@permission_required("products.edit")
def update_product(pid):
    existing = query("SELECT * FROM products WHERE id = ?", (pid,), one=True)
    if not existing:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(silent=True) or {}
    fields = _clean(data)
    barcode = fields.get("barcode")
    if barcode:
        dup = query("SELECT id FROM products WHERE barcode = ? AND id != ?", (barcode, pid), one=True)
        if dup:
            return jsonify({"error": "barcode already exists"}), 409

    merged = dict(existing)
    merged.update(fields)
    fields["search_blob"] = build_search_blob(
        merged.get("name_en", ""), merged.get("name_ar", ""),
        merged.get("name_fr", ""), merged.get("sku", ""), merged.get("barcode", "") or "",
    )
    set_sql = ", ".join(f"{c} = ?" for c in fields)
    execute(
        f"UPDATE products SET {set_sql}, updated_at = datetime('now','localtime') WHERE id = ?",
        [fields[c] for c in fields] + [pid],
    )
    # If quantity was directly edited, log the adjustment movement.
    if "quantity" in fields and float(fields["quantity"]) != float(existing["quantity"]):
        delta = float(fields["quantity"]) - float(existing["quantity"])
        _record_movement(pid, delta, float(fields["quantity"]), "adjustment", note="Manual edit")
    log_action("update_product", "product", pid, merged.get("name_en", ""))
    return jsonify({"ok": True})


@bp.post("/<int:pid>/archive")
@permission_required("products.edit")
def archive_product(pid):
    execute("UPDATE products SET is_archived = 1 WHERE id = ?", (pid,))
    log_action("archive_product", "product", pid)
    return jsonify({"ok": True})


@bp.post("/<int:pid>/restore")
@permission_required("products.edit")
def restore_product(pid):
    execute("UPDATE products SET is_archived = 0 WHERE id = ?", (pid,))
    log_action("restore_product", "product", pid)
    return jsonify({"ok": True})


@bp.delete("/<int:pid>")
@permission_required("products.edit")
def delete_product(pid):
    used = query("SELECT COUNT(*) AS c FROM sale_items WHERE product_id = ?", (pid,), one=True)["c"]
    if used:
        # Keep referential history — archive instead of hard delete.
        execute("UPDATE products SET is_archived = 1 WHERE id = ?", (pid,))
        log_action("archive_product", "product", pid, "archived (has sales history)")
        return jsonify({"ok": True, "archived": True})
    execute("DELETE FROM products WHERE id = ?", (pid,))
    log_action("delete_product", "product", pid)
    return jsonify({"ok": True})


@bp.post("/<int:pid>/duplicate")
@permission_required("products.edit")
def duplicate_product(pid):
    row = query("SELECT * FROM products WHERE id = ?", (pid,), one=True)
    if not row:
        return jsonify({"error": "not found"}), 404
    d = dict(row)
    d["name_en"] = (d["name_en"] or "") + " (copy)"
    d["barcode"] = None
    d["sku"] = (d["sku"] or "") + "-C"
    d["quantity"] = 0
    fields = _clean(d)
    fields["search_blob"] = build_search_blob(
        fields.get("name_en", ""), fields.get("name_ar", ""), fields.get("name_fr", ""),
    )
    cols = list(fields.keys())
    new_id = execute(
        f"INSERT INTO products ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
        [fields[c] for c in cols],
    )
    log_action("duplicate_product", "product", new_id, f"from #{pid}")
    return jsonify({"id": new_id}), 201


def _record_movement(product_id, change_qty, balance, reason, ref_type="", ref_id=None, note=""):
    from flask import session
    execute(
        "INSERT INTO inventory_movements "
        "(product_id, change_qty, balance, reason, ref_type, ref_id, note, user_id) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (product_id, change_qty, balance, reason, ref_type, ref_id, note,
         session.get("user_id")),
    )


# ---------------------------------------------------------------------------
# Excel / CSV import & export
# ---------------------------------------------------------------------------
IMPORT_HEADERS = [
    "barcode", "sku", "name_en", "name_ar", "name_fr", "category", "brand",
    "purchase_price", "selling_price", "tax_rate", "quantity", "min_stock", "unit",
]


@bp.get("/export")
@login_required
def export_products():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(IMPORT_HEADERS)
    rows = query(
        "SELECT p.*, c.name AS category, b.name AS brand FROM products p "
        "LEFT JOIN categories c ON c.id=p.category_id "
        "LEFT JOIN brands b ON b.id=p.brand_id WHERE p.is_archived=0"
    )
    for r in rows:
        ws.append([
            r["barcode"], r["sku"], r["name_en"], r["name_ar"], r["name_fr"],
            r["category"], r["brand"], r["purchase_price"], r["selling_price"],
            r["tax_rate"], r["quantity"], r["min_stock"], r["unit"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action("export_products", "product")
    return send_file(
        buf, as_attachment=True, download_name="products_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/import-template")
@login_required
def import_template():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(IMPORT_HEADERS)
    ws.append(["6111234567890", "SKU001", "Coca Cola 1L", "كوكا كولا",
               "Coca Cola 1L", "Beverages", "Coca Cola", 6.0, 9.0, 20, 100, 20, "pcs"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/import")
@permission_required("products.edit")
def import_products():
    from openpyxl import load_workbook

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "no file uploaded"}), 400
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return jsonify({"error": "invalid Excel file"}), 400
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "empty file"}), 400

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    created = updated = skipped = 0
    db = get_db()

    for raw in rows[1:]:
        if not raw or all(c is None for c in raw):
            continue
        def val(key, default=""):
            i = idx.get(key)
            return raw[i] if (i is not None and i < len(raw) and raw[i] is not None) else default

        name_en = str(val("name_en") or "").strip()
        name_ar = str(val("name_ar") or "").strip()
        if not (name_en or name_ar):
            skipped += 1
            continue
        barcode = str(val("barcode") or "").strip() or None
        cat_id = _lookup_or_create(db, "categories", "name", str(val("category") or "").strip())
        brand_id = _lookup_or_create(db, "brands", "name", str(val("brand") or "").strip())

        record = {
            "barcode": barcode,
            "sku": str(val("sku") or "").strip(),
            "name_en": name_en, "name_ar": name_ar,
            "name_fr": str(val("name_fr") or "").strip(),
            "category_id": cat_id, "brand_id": brand_id,
            "purchase_price": _num(val("purchase_price")),
            "selling_price": _num(val("selling_price")),
            "tax_rate": _num(val("tax_rate")),
            "quantity": _num(val("quantity")),
            "min_stock": _num(val("min_stock")),
            "unit": str(val("unit") or "pcs").strip() or "pcs",
        }
        record["search_blob"] = build_search_blob(
            name_en, name_ar, record["name_fr"], record["sku"], barcode or "")

        existing = None
        if barcode:
            existing = db.execute("SELECT id FROM products WHERE barcode=?", (barcode,)).fetchone()
        if existing:
            cols = [c for c in record if c != "barcode"]
            db.execute(
                f"UPDATE products SET {', '.join(f'{c}=?' for c in cols)} WHERE id=?",
                [record[c] for c in cols] + [existing["id"]],
            )
            updated += 1
        else:
            cols = list(record.keys())
            db.execute(
                f"INSERT INTO products ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
                [record[c] for c in cols],
            )
            created += 1
    db.commit()
    log_action("import_products", "product", detail=f"created={created} updated={updated}")
    return jsonify({"created": created, "updated": updated, "skipped": skipped})


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _lookup_or_create(db, table, col, value):
    if not value:
        return None
    row = db.execute(f"SELECT id FROM {table} WHERE {col}=?", (value,)).fetchone()
    if row:
        return row["id"]
    cur = db.execute(f"INSERT INTO {table} ({col}) VALUES (?)", (value,))
    return cur.lastrowid
